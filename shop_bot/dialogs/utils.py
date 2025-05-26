import logging
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


async def save_order_to_excel(
    order_data,
    file_path='orders.xlsx',
):
    file = Path(file_path)

    if file.exists():
        wb = load_workbook(file)
        ws = wb.active

    else:
        wb = Workbook()
        ws = wb.active
        ws.append(
            [
                'ID',
                'ФИО',
                'Телефон',
                'Город',
                'Улица',
                'Дом',
                'Квартира',
                'Товары',
                'Количество',
                'Общая сумма заказа',
            ]
        )

    last_row = ws.max_row
    order_id = (
        1 if last_row == 1 else ws.cell(row=last_row, column=1).value + 1
    )

    name, phone, city, street, house, flat, items, total_price = (
        order_data
    )

    for name_item, amount in items:
        ws.append(
            [
                order_id,
                name,
                phone,
                city,
                street,
                house,
                flat,
                name_item,
                amount,
                total_price,
            ]
        )

    wb.save(file)
    logger.info(f"Заказ №{order_id} успешно добавлен.")
